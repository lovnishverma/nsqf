"""
NSQF Exam Document Generator — NIELIT Ropar
Template-Based Injection strategy with full per-course template registry.
Supports all 12 NSQF courses with course-specific template sets.
"""

import io
import os
import re
import zipfile

import pandas as pd
from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "templates", "xlsx")

REQUIRED_COLUMNS = {
    "Course_name",
    "Module_Short_Name",
    "Candidate_registration_no",
    "Candidate_Name",
    "Candidate_Gender",
    "Father_Name",
    "Mother_Name",
    "Candidate_Birth_date",
    "Institute_Name",
    "Roll_Number",
    "Exam Date",
    "Exam Time",
}

# ---------------------------------------------------------------------------
# Course Registry
# Each course maps Module_Short_Name values to their document specs.
# "doc_type" drives which generator function is called.
# Supported doc_types:
#   theory_attendance   — theory paper attendance sheet
#   practical_attendance — practical exam attendance sheet
#   typing_attendance   — typing test sheet (I059 only)
#   internal_assessment — internal assessment marks sheet
#   compiled_result_hq  — compiled result for HQ submission
# ---------------------------------------------------------------------------

COURSE_REGISTRY = {
    # ── Single-module courses (1 Theory + 1 Practical) ─────────────────────
    "I091": {
        "label": "Essentials of AI",
        "modules": {
            "__all__": [
                {"doc_type": "theory_attendance",    "template": "I091_theory_attendance.xlsx",    "suffix": "Theory_Attendance"},
                {"doc_type": "practical_attendance", "template": "I091_practical_attendance.xlsx", "suffix": "Practical_Attendance"},
                {"doc_type": "internal_assessment",  "template": "I091_internal_assessment.xlsx",  "suffix": "Internal_Assessment"},
                {"doc_type": "compiled_result_hq",   "template": "I091_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"},
            ]
        },
    },
    "I092": {
        "label": "Essentials of Cloud Computing & Virtualisation",
        "modules": {
            "__all__": [
                {"doc_type": "theory_attendance",    "template": "I092_theory_attendance.xlsx",    "suffix": "Theory_Attendance"},
                {"doc_type": "practical_attendance", "template": "I092_practical_attendance.xlsx", "suffix": "Practical_Attendance"},
                {"doc_type": "internal_assessment",  "template": "I092_internal_assessment.xlsx",  "suffix": "Internal_Assessment"},
                {"doc_type": "compiled_result_hq",   "template": "I092_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"},
            ]
        },
    },
    "I081": {
        "label": "Fundamentals of Data Curation (Python)",
        "modules": {
            "__all__": [
                {"doc_type": "theory_attendance",    "template": "I081_theory_attendance.xlsx",    "suffix": "Theory_Attendance"},
                {"doc_type": "practical_attendance", "template": "I081_practical_attendance.xlsx", "suffix": "Practical_Attendance"},
                {"doc_type": "internal_assessment",  "template": "I081_internal_assessment.xlsx",  "suffix": "Internal_Assessment"},
                {"doc_type": "compiled_result_hq",   "template": "I081_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"},
            ]
        },
    },
    "I079": {
        "label": "Fundamentals of Data Annotation (Python)",
        "modules": {
            "__all__": [
                {"doc_type": "theory_attendance",    "template": "I079_theory_attendance.xlsx",    "suffix": "Theory_Attendance"},
                {"doc_type": "practical_attendance", "template": "I079_practical_attendance.xlsx", "suffix": "Practical_Attendance"},
                {"doc_type": "internal_assessment",  "template": "I079_internal_assessment.xlsx",  "suffix": "Internal_Assessment"},
                {"doc_type": "compiled_result_hq",   "template": "I079_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"},
            ]
        },
    },
    "I060": {
        "label": "Certified Computer Application, Accounting & Publishing",
        "modules": {
            "__all__": [
                {"doc_type": "theory_attendance",    "template": "I060_theory_attendance.xlsx",    "suffix": "Theory_Attendance"},
                {"doc_type": "practical_attendance", "template": "I060_practical_attendance.xlsx", "suffix": "Practical_Attendance"},
                {"doc_type": "internal_assessment",  "template": "I060_internal_assessment.xlsx",  "suffix": "Internal_Assessment"},
                {"doc_type": "compiled_result_hq",   "template": "I060_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"},
            ]
        },
    },

    # ── I059 — has additional Typing test sheet ─────────────────────────────
    "I059": {
        "label": "Certified Data Entry & Office Assistant",
        "modules": {
            "__all__": [
                {"doc_type": "theory_attendance",    "template": "I059_theory_attendance.xlsx",    "suffix": "Theory_Attendance"},
                {"doc_type": "practical_attendance", "template": "I059_practical_attendance.xlsx", "suffix": "Practical_Attendance"},
                {"doc_type": "typing_attendance",    "template": "I059_typing_attendance.xlsx",    "suffix": "Typing_Attendance"},
                {"doc_type": "internal_assessment",  "template": "I059_internal_assessment.xlsx",  "suffix": "Internal_Assessment"},
                {"doc_type": "compiled_result_hq",   "template": "I059_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"},
            ]
        },
    },

    # ── Two-Theory-module courses ────────────────────────────────────────────
    # Module_Short_Name from CSV drives which theory template is used.
    # "__practical__", "__internal__", "__compiled__" are special keys
    # for doc types that cover ALL modules combined (one file per course).
    "I090": {
        "label": "Full Stack Development Associate",
        "modules": {
            "M1": [
                {"doc_type": "theory_attendance", "template": "I090_theory1_attendance.xlsx", "suffix": "Theory1_Attendance"},
            ],
            "M2": [
                {"doc_type": "theory_attendance", "template": "I090_theory2_attendance.xlsx", "suffix": "Theory2_Attendance"},
            ],
            "__practical__": [
                {"doc_type": "practical_attendance", "template": "I090_practical_attendance.xlsx", "suffix": "Practical_Attendance"},
            ],
            "__internal__": [
                {"doc_type": "internal_assessment", "template": "I090_internal_assessment.xlsx", "suffix": "Internal_Assessment"},
            ],
            "__compiled__": [
                {"doc_type": "compiled_result_hq", "template": "I090_compiled_result_hq.xlsx", "suffix": "Compiled_Result_HQ"},
            ],
        },
    },
    "I077": {
        "label": "Multimedia Development Associate",
        "modules": {
            "M1": [{"doc_type": "theory_attendance", "template": "I077_theory1_attendance.xlsx", "suffix": "Theory1_Attendance"}],
            "M2": [{"doc_type": "theory_attendance", "template": "I077_theory2_attendance.xlsx", "suffix": "Theory2_Attendance"}],
            "__practical__": [{"doc_type": "practical_attendance", "template": "I077_practical_attendance.xlsx", "suffix": "Practical_Attendance"}],
            "__internal__":  [{"doc_type": "internal_assessment",  "template": "I077_internal_assessment.xlsx",  "suffix": "Internal_Assessment"}],
            "__compiled__":  [{"doc_type": "compiled_result_hq",   "template": "I077_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"}],
        },
    },
    "I078": {
        "label": "ITeS BPO Executive - Voice",
        "modules": {
            "M1": [{"doc_type": "theory_attendance", "template": "I078_theory1_attendance.xlsx", "suffix": "Theory1_Attendance"}],
            "M2": [{"doc_type": "theory_attendance", "template": "I078_theory2_attendance.xlsx", "suffix": "Theory2_Attendance"}],
            "__practical__": [{"doc_type": "practical_attendance", "template": "I078_practical_attendance.xlsx", "suffix": "Practical_Attendance"}],
            "__internal__":  [{"doc_type": "internal_assessment",  "template": "I078_internal_assessment.xlsx",  "suffix": "Internal_Assessment"}],
            "__compiled__":  [{"doc_type": "compiled_result_hq",   "template": "I078_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"}],
        },
    },
    "E063": {
        "label": "Assistant Computer Technician",
        "modules": {
            "M1": [{"doc_type": "theory_attendance", "template": "E063_theory1_attendance.xlsx", "suffix": "Theory1_Attendance"}],
            "M2": [{"doc_type": "theory_attendance", "template": "E063_theory2_attendance.xlsx", "suffix": "Theory2_Attendance"}],
            "__practical__": [{"doc_type": "practical_attendance", "template": "E063_practical_attendance.xlsx", "suffix": "Practical_Attendance"}],
            "__internal__":  [{"doc_type": "internal_assessment",  "template": "E063_internal_assessment.xlsx",  "suffix": "Internal_Assessment"}],
            "__compiled__":  [{"doc_type": "compiled_result_hq",   "template": "E063_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"}],
        },
    },
    "E058": {
        "label": "IoT Assistant",
        "modules": {
            "M1": [{"doc_type": "theory_attendance", "template": "E058_theory1_attendance.xlsx", "suffix": "Theory1_Attendance"}],
            "M2": [{"doc_type": "theory_attendance", "template": "E058_theory2_attendance.xlsx", "suffix": "Theory2_Attendance"}],
            "__practical__": [{"doc_type": "practical_attendance", "template": "E058_practical_attendance.xlsx", "suffix": "Practical_Attendance"}],
            "__internal__":  [{"doc_type": "internal_assessment",  "template": "E058_internal_assessment.xlsx",  "suffix": "Internal_Assessment"}],
            "__compiled__":  [{"doc_type": "compiled_result_hq",   "template": "E058_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"}],
        },
    },
    "E059": {
        "label": "IoT Associate",
        "modules": {
            "M1": [{"doc_type": "theory_attendance", "template": "E059_theory1_attendance.xlsx", "suffix": "Theory1_Attendance"}],
            "M2": [{"doc_type": "theory_attendance", "template": "E059_theory2_attendance.xlsx", "suffix": "Theory2_Attendance"}],
            "__practical__": [{"doc_type": "practical_attendance", "template": "E059_practical_attendance.xlsx", "suffix": "Practical_Attendance"}],
            "__internal__":  [{"doc_type": "internal_assessment",  "template": "E059_internal_assessment.xlsx",  "suffix": "Internal_Assessment"}],
            "__compiled__":  [{"doc_type": "compiled_result_hq",   "template": "E059_compiled_result_hq.xlsx",   "suffix": "Compiled_Result_HQ"}],
        },
    },
}

# Build a lookup: course_id → course entry, also supporting full Course_name strings
# so the CSV can have either the code or the label in the Course_name column.
_COURSE_BY_NAME: dict[str, str] = {}
for _cid, _cinfo in COURSE_REGISTRY.items():
    _COURSE_BY_NAME[_cid.upper()] = _cid
    _COURSE_BY_NAME[_cinfo["label"].upper()] = _cid


def resolve_course_id(course_name: str) -> str | None:
    """Return registry key for a given Course_name value, or None if unknown."""
    return _COURSE_BY_NAME.get(str(course_name).strip().upper())


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", str(name)).strip()


def _load_template(template_name: str):
    path = os.path.join(TEMPLATE_DIR, template_name)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Template not found: {path}")
    return load_workbook(path)


def _safe_val(value):
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _gender_initial(value) -> str:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(value).strip()
    return s[0].upper() if s else ""


# ---------------------------------------------------------------------------
# Core injection engines
# One function per doc_type; template filename is passed in as argument.
# ---------------------------------------------------------------------------

def _inject_theory_attendance(df: pd.DataFrame, template: str, institute: str, course: str) -> io.BytesIO:
    """
    Header: B1=Institute, B2=Course, H3=Exam Cycle
    Data starts row 5: A=SNo, B=Roll, C=RegNo, D=Name, E=Gender,
                        F=Father, G=Mother, H=DOB, I=ExamDate, J=ExamTime
    """
    wb = _load_template(template)
    ws = wb.active
    ws["B1"] = institute
    ws["B2"] = course
    dates = df["Exam Date"].dropna().unique()
    ws["H3"] = dates[0] if len(dates) else ""
    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = 4 + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, _safe_val(row["Roll_Number"]))
        ws.cell(r, 3, _safe_val(row["Candidate_registration_no"]))
        ws.cell(r, 4, _safe_val(row["Candidate_Name"]))
        ws.cell(r, 5, _gender_initial(row["Candidate_Gender"]))
        ws.cell(r, 6, _safe_val(row["Father_Name"]))
        ws.cell(r, 7, _safe_val(row["Mother_Name"]))
        ws.cell(r, 8, _safe_val(row["Candidate_Birth_date"]))
        ws.cell(r, 9, _safe_val(row["Exam Date"]))
        ws.cell(r, 10, _safe_val(row["Exam Time"]))
    return _save_wb(wb)


def _inject_practical_attendance(df: pd.DataFrame, template: str, institute: str, course: str) -> io.BytesIO:
    """
    Header: B1=Institute, B2=Course, H3=Exam Cycle
    Data starts row 8. Cols K-N blank (signatures/marks).
    """
    wb = _load_template(template)
    ws = wb.active
    ws["B1"] = institute
    ws["B2"] = course
    dates = df["Exam Date"].dropna().unique()
    ws["H3"] = dates[0] if len(dates) else ""
    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = 8 + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, _safe_val(row["Roll_Number"]))
        ws.cell(r, 3, _safe_val(row["Candidate_registration_no"]))
        ws.cell(r, 4, _safe_val(row["Candidate_Name"]))
        ws.cell(r, 5, _gender_initial(row["Candidate_Gender"]))
        ws.cell(r, 6, _safe_val(row["Father_Name"]))
        ws.cell(r, 7, _safe_val(row["Mother_Name"]))
        ws.cell(r, 8, _safe_val(row["Candidate_Birth_date"]))
        # ws.cell(r, 9, _safe_val(row["Exam Date"]))
        # ws.cell(r, 10, _safe_val(row["Exam Time"]))
        # Cols 11-14 (K-N) intentionally blank — physical signature columns
    return _save_wb(wb)


def _inject_typing_attendance(df: pd.DataFrame, template: str, institute: str, course: str) -> io.BytesIO:
    """
    I059 only — Typing Test Attendance Sheet.
    Header: B1=Institute, B2=Course, H3=Exam Cycle
    Data starts row 5: A=SNo, B=Roll, C=RegNo, D=Name, E=Gender,
                        F=Father, G=DOB, H=ExamDate, I=ExamTime
    Cols J-L blank (speed/accuracy/result — filled physically).
    """
    wb = _load_template(template)
    ws = wb.active
    ws["B1"] = institute
    ws["B2"] = course
    dates = df["Exam Date"].dropna().unique()
    ws["H3"] = dates[0] if len(dates) else ""
    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = 4 + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, _safe_val(row["Roll_Number"]))
        ws.cell(r, 3, _safe_val(row["Candidate_registration_no"]))
        ws.cell(r, 4, _safe_val(row["Candidate_Name"]))
        ws.cell(r, 5, _gender_initial(row["Candidate_Gender"]))
        ws.cell(r, 6, _safe_val(row["Father_Name"]))
        ws.cell(r, 7, _safe_val(row["Candidate_Birth_date"]))
        ws.cell(r, 8, _safe_val(row["Exam Date"]))
        ws.cell(r, 9, _safe_val(row["Exam Time"]))
        # Cols 10-12 (J-L) blank — typing speed/accuracy filled physically
    return _save_wb(wb)


def _inject_internal_assessment(df: pd.DataFrame, template: str, institute: str, course: str) -> io.BytesIO:
    """
    Header: B1=Institute, B2=Course
    Data starts row 5: A=SNo, B=Roll, C=RegNo, D=Name, E=Gender, F=Father, G=DOB
    """
    wb = _load_template(template)
    ws = wb.active
    ws["B1"] = institute
    ws["B2"] = course
    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = 4 + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, _safe_val(row["Roll_Number"]))
        ws.cell(r, 3, _safe_val(row["Candidate_registration_no"]))
        ws.cell(r, 4, _safe_val(row["Candidate_Name"]))
        ws.cell(r, 5, _gender_initial(row["Candidate_Gender"]))
        ws.cell(r, 6, _safe_val(row["Father_Name"]))
        ws.cell(r, 7, _safe_val(row["Candidate_Birth_date"]))
    return _save_wb(wb)


def _inject_compiled_result_hq(df: pd.DataFrame, template: str, institute: str, course: str) -> io.BytesIO:
    """
    Header: D2=Institute, D4=Course
    Data starts row 9: A=SNo, B=Roll, C=RegNo, D=Name, E=Gender,
                        F=Father, G=Mother, H=DOB
    """
    wb = _load_template(template)
    ws = wb.active
    ws["D2"] = institute
    ws["D4"] = course
    for i, (_, row) in enumerate(df.iterrows(), 1):
        r = 8 + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, _safe_val(row["Roll_Number"]))
        ws.cell(r, 3, _safe_val(row["Candidate_registration_no"]))
        ws.cell(r, 4, _safe_val(row["Candidate_Name"]))
        ws.cell(r, 5, _gender_initial(row["Candidate_Gender"]))
        ws.cell(r, 6, _safe_val(row["Father_Name"]))
        ws.cell(r, 7, _safe_val(row["Mother_Name"]))
        ws.cell(r, 8, _safe_val(row["Candidate_Birth_date"]))
    return _save_wb(wb)


# Map doc_type string → injection function
DOC_INJECTORS = {
    "theory_attendance":    _inject_theory_attendance,
    "practical_attendance": _inject_practical_attendance,
    "typing_attendance":    _inject_typing_attendance,
    "internal_assessment":  _inject_internal_assessment,
    "compiled_result_hq":   _inject_compiled_result_hq,
}


def _save_wb(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Per-course generation orchestrator
# ---------------------------------------------------------------------------

def generate_docs_for_course(
    course_df: pd.DataFrame,
    course_id: str,
    institute: str,
    course_name: str,
    zf: zipfile.ZipFile,
    safe_inst: str,
    safe_course: str,
) -> int:
    """
    Generate all documents for one (institute, course) pair.
    Returns count of files written.
    """
    registry = COURSE_REGISTRY[course_id]
    modules_map = registry["modules"]
    files_written = 0

    # ── Single-module courses (key "__all__") ────────────────────────────────
    if "__all__" in modules_map:
        specs = modules_map["__all__"]
        for spec in specs:
            files_written += _run_spec(
                spec, course_df, institute, course_name,
                safe_inst, safe_course, zf
            )
        return files_written

    # ── Multi-module courses ─────────────────────────────────────────────────
    # 1. Theory: group by Module_Short_Name and match M1/M2/etc.
    for module_key, specs in modules_map.items():
        if module_key.startswith("__"):
            continue  # handled separately below
        module_df = course_df[course_df["Module_Short_Name"].str.strip().str.upper() == module_key.upper()]
        if module_df.empty:
            continue
        for spec in specs:
            suffix_with_module = f"{module_key}_{spec['suffix']}"
            files_written += _run_spec(
                spec, module_df, institute, course_name,
                safe_inst, safe_course, zf,
                suffix_override=suffix_with_module
            )

    # 2. Practical / Internal / Compiled — use full course_df (all modules together)
    for special_key in ("__practical__", "__internal__", "__compiled__"):
        if special_key not in modules_map:
            continue
        for spec in modules_map[special_key]:
            files_written += _run_spec(
                spec, course_df, institute, course_name,
                safe_inst, safe_course, zf
            )

    return files_written


def _run_spec(
    spec: dict,
    df: pd.DataFrame,
    institute: str,
    course_name: str,
    safe_inst: str,
    safe_course: str,
    zf: zipfile.ZipFile,
    suffix_override: str | None = None,
) -> int:
    suffix = suffix_override or spec["suffix"]
    doc_type = spec["doc_type"]
    template = spec["template"]
    injector = DOC_INJECTORS.get(doc_type)
    if injector is None:
        return 0
    try:
        buf = injector(df, template, institute, course_name)
        filename = f"{safe_inst}_{safe_course}_{suffix}.xlsx"
        arcname = os.path.join(safe_inst, filename)
        zf.writestr(arcname, buf.read())
        return 1
    except FileNotFoundError as e:
        note = f"SKIPPED — template missing: {e}\n"
        zf.writestr(os.path.join(safe_inst, f"MISSING_{suffix}.txt"), note)
        return 0
    except Exception as e:
        note = f"ERROR generating {suffix} for {institute} / {course_name}:\n{e}\n"
        zf.writestr(os.path.join(safe_inst, f"ERROR_{suffix}.txt"), note)
        return 0


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    courses_for_ui = [
        {"id": cid, "label": info["label"]}
        for cid, info in COURSE_REGISTRY.items()
    ]
    return render_template("index.html", courses=courses_for_ui)


@app.route("/generate", methods=["POST"])
def generate():
    if "csv_file" not in request.files or request.files["csv_file"].filename == "":
        return jsonify({"error": "No CSV file uploaded."}), 400

    try:
        df = pd.read_csv(request.files["csv_file"], dtype=str)
    except Exception as e:
        return jsonify({"error": f"Failed to parse CSV: {e}"}), 400

    df.columns = df.columns.str.strip()

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        return jsonify({
            "error": "CSV is missing required columns.",
            "missing_columns": sorted(missing),
        }), 400

    for col in ["Institute_Name", "Course_name", "Module_Short_Name"]:
        df[col] = df[col].str.strip()

    # Attach resolved course_id column
    df["_course_id"] = df["Course_name"].apply(resolve_course_id)

    unknown = df[df["_course_id"].isna()]["Course_name"].unique().tolist()
    known_df = df[df["_course_id"].notna()]

    if known_df.empty:
        return jsonify({
            "error": "No recognised courses found in CSV.",
            "unrecognised_courses": unknown,
        }), 400

    zip_buf = io.BytesIO()
    files_written = 0

    try:
        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            # Log unrecognised courses as a README in the ZIP root
            if unknown:
                note = "The following Course_name values were not recognised and were skipped:\n"
                note += "\n".join(f"  - {c}" for c in unknown)
                zf.writestr("UNRECOGNISED_COURSES.txt", note)

            for (institute, course_name, course_id), group_df in known_df.groupby(
                ["Institute_Name", "Course_name", "_course_id"], sort=True
            ):
                if group_df.empty:
                    continue
                safe_inst = sanitize_filename(institute)
                safe_course = sanitize_filename(course_name)
                files_written += generate_docs_for_course(
                    group_df, course_id, institute, course_name,
                    zf, safe_inst, safe_course
                )

    except Exception as e:
        return jsonify({"error": f"Failed to build ZIP: {e}"}), 500

    if files_written == 0:
        return jsonify({
            "error": "No documents were generated. Verify that templates exist in templates/xlsx/ and the CSV has valid data.",
        }), 500

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name="NSQF_Documents.zip",
    )


@app.route("/courses")
def list_courses():
    """JSON endpoint listing all registered courses — useful for debugging."""
    return jsonify({
        cid: {
            "label": info["label"],
            "module_keys": list(info["modules"].keys()),
        }
        for cid, info in COURSE_REGISTRY.items()
    })


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True, port=5000)
